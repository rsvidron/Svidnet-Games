"""
Ranked lists, admin collections, and friends rankings endpoints.
"""
from fastapi import APIRouter, Depends, HTTPException, Header, Query
from sqlalchemy.orm import Session
from sqlalchemy import asc, desc
from datetime import datetime, timezone
from typing import Optional, List
from pydantic import BaseModel
from jose import jwt, JWTError
import httpx
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from models.ranked_list import RankedList, RankedListItem
from models.collection import Collection, CollectionItem
from models.friends import Friendship
from models.user import User

router = APIRouter(prefix="/api/rankings", tags=["rankings"])

TMDB_BEARER = os.getenv(
    "TMDB_BEARER_TOKEN",
    "eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiI2ZjRlMGZhZWJjZTY3Yjk4NzYxMmIxZmZjZmI3ODg1ZSIsIm5iZiI6MTcyMjQ0MzkzMi43OTcsInN1YiI6IjY2YWE2ODljNTliYWI3OGRiNzBhZjhiNSIsInNjb3BlcyI6WyJhcGlfcmVhZCJdLCJ2ZXJzaW9uIjoxfQ.4Y4Lu7P-vaFpK07YlMutOHD5yuk4iATpWF28K4yapTQ"
)
TMDB_BASE = "https://api.themoviedb.org/3"
TMDB_HEADERS = {"Authorization": f"Bearer {TMDB_BEARER}", "accept": "application/json"}

ADMIN_USERNAMES = ["svidthekid"]
ADMIN_EMAILS = ["svidron.robert@gmail.com"]


# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    from database import SessionLocal
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ── Auth ──────────────────────────────────────────────────────────────────────
def get_current_user_id(authorization: Optional[str] = Header(None)) -> int:
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.removeprefix("Bearer ").strip()
    secret = os.getenv("SECRET_KEY", "test-secret-key-for-development")
    try:
        payload = jwt.decode(token, secret, algorithms=["HS256"])
        user_id = payload.get("sub")
        if user_id is None:
            user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token claims")
        return int(user_id)
    except HTTPException:
        raise
    except JWTError as e:
        print(f"[rankings] JWT error: {e}")
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    except Exception as e:
        print(f"[rankings] Auth error: {type(e).__name__}: {e}")
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def _require_admin(user_id: int, db: Session):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=403, detail="Forbidden")
    if not (user.role == "admin" or user.username in ADMIN_USERNAMES or user.email in ADMIN_EMAILS):
        raise HTTPException(status_code=403, detail="Admin only")


# ── Schemas ───────────────────────────────────────────────────────────────────
class ListCreate(BaseModel):
    title: str
    description: Optional[str] = None
    is_public: bool = False

class ListUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    is_public: Optional[bool] = None

class ItemAdd(BaseModel):
    media_id: str
    media_type: str
    title: str
    poster_path: Optional[str] = None
    release_year: Optional[str] = None
    overview: Optional[str] = None
    note: Optional[str] = None

class ItemUpdate(BaseModel):
    note: Optional[str] = None

class ReorderBody(BaseModel):
    item_ids: List[int]

class CollectionCreate(BaseModel):
    title: str
    description: Optional[str] = None
    cover_poster: Optional[str] = None

class CollectionUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    cover_poster: Optional[str] = None
    is_active: Optional[bool] = None

class CollectionItemAdd(BaseModel):
    media_id: str
    media_type: str
    title: str
    poster_path: Optional[str] = None
    release_year: Optional[str] = None
    overview: Optional[str] = None
    sort_order: int = 0


# ═══════════════════════════════════════════════════════════════════════════════
# USER RANKED LISTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/my")
def get_my_lists(
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    lists = db.query(RankedList).filter(RankedList.user_id == user_id).order_by(desc(RankedList.updated_at)).all()
    return [_list_dict(lst, db) for lst in lists]


@router.post("/my")
def create_list(
    body: ListCreate,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    lst = RankedList(user_id=user_id, title=body.title, description=body.description, is_public=body.is_public)
    db.add(lst)
    db.commit()
    db.refresh(lst)
    return _list_dict(lst, db)


@router.get("/my/{list_id}")
def get_list(
    list_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    lst = _get_owned_list(list_id, user_id, db)
    items = db.query(RankedListItem).filter(RankedListItem.list_id == list_id).order_by(asc(RankedListItem.rank)).all()
    result = _list_dict(lst, db)
    result["items"] = [_item_dict(i) for i in items]
    return result


@router.patch("/my/{list_id}")
def update_list(
    list_id: int,
    body: ListUpdate,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    lst = _get_owned_list(list_id, user_id, db)
    if body.title is not None:
        lst.title = body.title
    if body.description is not None:
        lst.description = body.description
    if body.is_public is not None:
        lst.is_public = body.is_public
    lst.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(lst)
    return _list_dict(lst, db)


@router.delete("/my/{list_id}")
def delete_list(
    list_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    lst = _get_owned_list(list_id, user_id, db)
    db.delete(lst)
    db.commit()
    return {"ok": True}


@router.post("/my/{list_id}/items")
def add_item(
    list_id: int,
    body: ItemAdd,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _get_owned_list(list_id, user_id, db)
    existing = db.query(RankedListItem).filter(
        RankedListItem.list_id == list_id,
        RankedListItem.media_id == body.media_id,
        RankedListItem.media_type == body.media_type,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Already in this list")
    max_rank = db.query(RankedListItem).filter(RankedListItem.list_id == list_id).count()
    item = RankedListItem(
        list_id=list_id, rank=max_rank + 1,
        media_id=body.media_id, media_type=body.media_type, title=body.title,
        poster_path=body.poster_path, release_year=body.release_year,
        overview=body.overview, note=body.note,
    )
    db.add(item)
    lst = db.query(RankedList).filter(RankedList.id == list_id).first()
    lst.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(item)
    return _item_dict(item)


@router.patch("/my/{list_id}/items/{item_id}")
def update_item(
    list_id: int, item_id: int, body: ItemUpdate,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _get_owned_list(list_id, user_id, db)
    item = db.query(RankedListItem).filter(RankedListItem.id == item_id, RankedListItem.list_id == list_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    if body.note is not None:
        item.note = body.note
    db.commit()
    db.refresh(item)
    return _item_dict(item)


@router.delete("/my/{list_id}/items/{item_id}")
def remove_item(
    list_id: int, item_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _get_owned_list(list_id, user_id, db)
    item = db.query(RankedListItem).filter(RankedListItem.id == item_id, RankedListItem.list_id == list_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    removed_rank = item.rank
    db.delete(item)
    for r in db.query(RankedListItem).filter(RankedListItem.list_id == list_id, RankedListItem.rank > removed_rank).all():
        r.rank -= 1
    db.query(RankedList).filter(RankedList.id == list_id).first().updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"ok": True}


@router.put("/my/{list_id}/reorder")
def reorder_items(
    list_id: int, body: ReorderBody,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _get_owned_list(list_id, user_id, db)
    items = db.query(RankedListItem).filter(RankedListItem.list_id == list_id).all()
    item_map = {i.id: i for i in items}
    if set(body.item_ids) != set(item_map.keys()):
        raise HTTPException(status_code=400, detail="item_ids must contain all items in the list")
    for rank, item_id in enumerate(body.item_ids, start=1):
        item_map[item_id].rank = rank
    db.query(RankedList).filter(RankedList.id == list_id).first().updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"ok": True}


# ── Clone a collection into user's ranked list ─────────────────────────────
@router.post("/my/from-collection/{collection_id}")
def create_list_from_collection(
    collection_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    """Create a new personal ranked list pre-populated from an admin collection."""
    col = db.query(Collection).filter(Collection.id == collection_id, Collection.is_active == True).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")

    # Check user doesn't already have a list from this collection
    existing = db.query(RankedList).filter(
        RankedList.user_id == user_id,
        RankedList.title == col.title,
    ).first()
    if existing:
        # Just return the existing one
        items = db.query(RankedListItem).filter(RankedListItem.list_id == existing.id).order_by(asc(RankedListItem.rank)).all()
        result = _list_dict(existing, db)
        result["items"] = [_item_dict(i) for i in items]
        return result

    lst = RankedList(user_id=user_id, title=col.title, description=col.description, is_public=False)
    db.add(lst)
    db.flush()

    col_items = db.query(CollectionItem).filter(CollectionItem.collection_id == collection_id).order_by(asc(CollectionItem.sort_order)).all()
    for idx, ci in enumerate(col_items, start=1):
        db.add(RankedListItem(
            list_id=lst.id, rank=idx,
            media_id=ci.media_id, media_type=ci.media_type, title=ci.title,
            poster_path=ci.poster_path, release_year=ci.release_year, overview=ci.overview,
        ))

    db.commit()
    db.refresh(lst)
    items = db.query(RankedListItem).filter(RankedListItem.list_id == lst.id).order_by(asc(RankedListItem.rank)).all()
    result = _list_dict(lst, db)
    result["items"] = [_item_dict(i) for i in items]
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN COLLECTIONS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/collections")
def list_collections(db: Session = Depends(get_db)):
    """Public: list all active admin collections."""
    cols = db.query(Collection).filter(Collection.is_active == True).order_by(asc(Collection.title)).all()
    return [_collection_dict(c, db) for c in cols]


@router.get("/collections/{collection_id}")
def get_collection(collection_id: int, db: Session = Depends(get_db)):
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")
    items = db.query(CollectionItem).filter(CollectionItem.collection_id == collection_id).order_by(asc(CollectionItem.sort_order)).all()
    result = _collection_dict(col, db)
    result["items"] = [_col_item_dict(i) for i in items]
    return result


@router.post("/collections")
def create_collection(
    body: CollectionCreate,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _require_admin(user_id, db)
    col = Collection(title=body.title, description=body.description, cover_poster=body.cover_poster, created_by=user_id)
    db.add(col)
    db.commit()
    db.refresh(col)
    return _collection_dict(col, db)


@router.patch("/collections/{collection_id}")
def update_collection(
    collection_id: int, body: CollectionUpdate,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _require_admin(user_id, db)
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")
    if body.title is not None:
        col.title = body.title
    if body.description is not None:
        col.description = body.description
    if body.cover_poster is not None:
        col.cover_poster = body.cover_poster
    if body.is_active is not None:
        col.is_active = body.is_active
    col.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(col)
    return _collection_dict(col, db)


@router.delete("/collections/{collection_id}")
def delete_collection(
    collection_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _require_admin(user_id, db)
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")
    db.delete(col)
    db.commit()
    return {"ok": True}


@router.post("/collections/{collection_id}/items")
def add_collection_item(
    collection_id: int, body: CollectionItemAdd,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _require_admin(user_id, db)
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")
    existing = db.query(CollectionItem).filter(
        CollectionItem.collection_id == collection_id,
        CollectionItem.media_id == body.media_id,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Already in this collection")
    item = CollectionItem(
        collection_id=collection_id, sort_order=body.sort_order,
        media_id=body.media_id, media_type=body.media_type, title=body.title,
        poster_path=body.poster_path, release_year=body.release_year, overview=body.overview,
    )
    db.add(item)
    col.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(item)
    return _col_item_dict(item)


@router.delete("/collections/{collection_id}/items/{item_id}")
def remove_collection_item(
    collection_id: int, item_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    _require_admin(user_id, db)
    item = db.query(CollectionItem).filter(CollectionItem.id == item_id, CollectionItem.collection_id == collection_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════════
# COLLECTION LEADERBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/collections/{collection_id}/leaderboard")
def get_collection_leaderboard(
    collection_id: int,
    db: Session = Depends(get_db),
):
    """
    For each item in the collection, compute the average rank across all users
    who have a ranked list matching this collection's title.
    Returns items sorted by average rank (lowest avg = best consensus rank).
    """
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="Collection not found")

    # Find all ranked lists whose title matches the collection title
    matching_lists = db.query(RankedList).filter(RankedList.title == col.title).all()
    if not matching_lists:
        return {"collection_id": collection_id, "title": col.title, "ranker_count": 0, "items": []}

    list_ids = [l.id for l in matching_lists]
    ranker_count = len(list_ids)

    # Fetch canonical items for reference (title, poster_path, media_type, release_year)
    canonical_items = db.query(CollectionItem).filter(
        CollectionItem.collection_id == collection_id
    ).order_by(asc(CollectionItem.sort_order)).all()

    # Build a lookup: title -> canonical info
    canonical_map = {ci.title: ci for ci in canonical_items}

    # Build list_id → user lookup
    list_user_map = {}
    for lst in matching_lists:
        user = db.query(User).filter(User.id == lst.user_id).first()
        list_user_map[lst.id] = user

    # Gather all ranked list items across all matching lists
    all_items = db.query(RankedListItem).filter(
        RankedListItem.list_id.in_(list_ids)
    ).all()

    # Accumulate per-user rank data per title
    rank_accumulator: dict[str, list[dict]] = {}
    for item in all_items:
        title = item.title
        user = list_user_map.get(item.list_id)
        if title not in rank_accumulator:
            rank_accumulator[title] = []
        rank_accumulator[title].append({
            "username": user.username if user else "Unknown",
            "avatar_url": user.avatar_url if user else None,
            "rank": item.rank,
        })

    # Build result: only include titles that appear in the canonical collection
    result_items = []
    for title, user_rankings in rank_accumulator.items():
        if title not in canonical_map:
            continue
        ci = canonical_map[title]
        avg_rank = sum(r["rank"] for r in user_rankings) / len(user_rankings)
        result_items.append({
            "title": title,
            "media_type": ci.media_type,
            "poster_path": ci.poster_path,
            "release_year": ci.release_year,
            "avg_rank": round(avg_rank, 2),
            "rank_count": len(user_rankings),
            "user_rankings": sorted(user_rankings, key=lambda x: x["rank"]),
        })

    # Sort by average rank ascending (best consensus position first)
    result_items.sort(key=lambda x: x["avg_rank"])

    # Re-number consensus positions 1..N
    for idx, item in enumerate(result_items, start=1):
        item["consensus_rank"] = idx

    return {
        "collection_id": collection_id,
        "title": col.title,
        "ranker_count": ranker_count,
        "items": result_items,
    }


@router.get("/collections/{collection_id}/leaderboard/export")
def export_collection_leaderboard(
    collection_id: int,
    db: Session = Depends(get_db),
):
    """
    Export leaderboard as an .xlsx file with two sheets:
      Sheet 1 – Consensus Leaderboard (avg rank, all user ranks per item)
      Sheet 2 – All Submissions (one column per user, rows = items in their ranked order)
    """
    import io
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — cannot generate Excel file")

    from fastapi.responses import StreamingResponse

    # ── Reuse the leaderboard query logic ─────────────────────────────────────
    col = db.query(Collection).filter(Collection.id == collection_id).first()
    if not col:
        raise HTTPException(404, "Collection not found")

    matching_lists = db.query(RankedList).filter(RankedList.title == col.title).all()
    if not matching_lists:
        raise HTTPException(404, "No submissions found for this collection")

    list_ids = [l.id for l in matching_lists]
    ranker_count = len(list_ids)

    canonical_items = db.query(CollectionItem).filter(
        CollectionItem.collection_id == collection_id
    ).order_by(asc(CollectionItem.sort_order)).all()
    canonical_map = {ci.title: ci for ci in canonical_items}

    list_user_map = {}
    for lst in matching_lists:
        user = db.query(User).filter(User.id == lst.user_id).first()
        list_user_map[lst.id] = user

    all_items = db.query(RankedListItem).filter(
        RankedListItem.list_id.in_(list_ids)
    ).all()

    rank_accumulator: dict[str, list[dict]] = {}
    for item in all_items:
        title = item.title
        user = list_user_map.get(item.list_id)
        if title not in rank_accumulator:
            rank_accumulator[title] = []
        rank_accumulator[title].append({
            "username": user.username if user else "Unknown",
            "rank": item.rank,
        })

    result_items = []
    for title, user_rankings in rank_accumulator.items():
        if title not in canonical_map:
            continue
        ci = canonical_map[title]
        avg_rank = sum(r["rank"] for r in user_rankings) / len(user_rankings)
        result_items.append({
            "title": title,
            "media_type": ci.media_type,
            "release_year": ci.release_year or "",
            "avg_rank": round(avg_rank, 2),
            "rank_count": len(user_rankings),
            "user_rankings": sorted(user_rankings, key=lambda x: x["rank"]),
        })
    result_items.sort(key=lambda x: x["avg_rank"])
    for idx, item in enumerate(result_items, start=1):
        item["consensus_rank"] = idx

    # ── Build Excel workbook ───────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Consensus Leaderboard ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Consensus Leaderboard"

    # Collect all usernames in submission order
    all_usernames = []
    seen = set()
    for item in result_items:
        for ur in item["user_rankings"]:
            if ur["username"] not in seen:
                all_usernames.append(ur["username"])
                seen.add(ur["username"])

    headers1 = ["Rank", "Title", "Type", "Year", "Avg Rank", "# Rankers"] + all_usernames
    for col_idx, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col_idx, value=h)

    for row_idx, item in enumerate(result_items, 2):
        user_rank_map = {ur["username"]: ur["rank"] for ur in item["user_rankings"]}
        row_data = [
            item["consensus_rank"],
            item["title"],
            item["media_type"].upper(),
            item["release_year"],
            item["avg_rank"],
            item["rank_count"],
        ] + [user_rank_map.get(u, "") for u in all_usernames]
        for col_idx, val in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val)

    # ── Sheet 2: All Submissions ───────────────────────────────────────────────
    ws2 = wb.create_sheet("All Submissions")

    # Build per-user submission: username -> sorted list of (rank, title)
    user_submissions: dict[str, list] = {}
    for item in all_items:
        user = list_user_map.get(item.list_id)
        uname = user.username if user else "Unknown"
        if uname not in user_submissions:
            user_submissions[uname] = []
        user_submissions[uname].append((item.rank, item.title))

    for uname in user_submissions:
        user_submissions[uname].sort(key=lambda x: x[0])

    usernames_sorted = sorted(user_submissions.keys())

    # Row 1: usernames as column headers
    for col_idx, uname in enumerate(usernames_sorted, 1):
        ws2.cell(row=1, column=col_idx, value=uname)

    # Rows 2+: each user's ranked items in order
    for col_idx, uname in enumerate(usernames_sorted, 1):
        for rank_idx, (rank, title) in enumerate(user_submissions[uname], 2):
            ws2.cell(row=rank_idx, column=col_idx, value=f"#{rank} {title}")

    # ── Stream response ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in col.title)[:40]
    filename = f"{safe_title}_leaderboard.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# FRIENDS RANKINGS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/friends")
def get_friends_rankings(
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    """Return all public ranked lists from the current user's friends."""
    # Collect friend IDs (friendship is bidirectional — stored as two rows)
    friend_ids = set()
    for f in db.query(Friendship).filter(Friendship.user_id == user_id).all():
        friend_ids.add(f.friend_id)
    for f in db.query(Friendship).filter(Friendship.friend_id == user_id).all():
        friend_ids.add(f.user_id)

    if not friend_ids:
        return []

    lists = db.query(RankedList).filter(
        RankedList.user_id.in_(friend_ids),
        RankedList.is_public == True,
    ).order_by(desc(RankedList.updated_at)).all()

    result = []
    for lst in lists:
        owner = db.query(User).filter(User.id == lst.user_id).first()
        d = _list_dict(lst, db)
        d["owner"] = {"id": owner.id, "username": owner.username, "avatar_url": owner.avatar_url} if owner else None
        result.append(d)
    return result


@router.get("/friends/{friend_user_id}/lists/{list_id}")
def get_friend_list_detail(
    friend_user_id: int, list_id: int,
    user_id: int = Depends(get_current_user_id),
    db: Session = Depends(get_db),
):
    """Return a friend's public ranked list with all items."""
    # Verify friendship
    is_friend = (
        db.query(Friendship).filter(Friendship.user_id == user_id, Friendship.friend_id == friend_user_id).first() or
        db.query(Friendship).filter(Friendship.user_id == friend_user_id, Friendship.friend_id == user_id).first()
    )
    if not is_friend:
        raise HTTPException(status_code=403, detail="Not friends")

    lst = db.query(RankedList).filter(RankedList.id == list_id, RankedList.user_id == friend_user_id, RankedList.is_public == True).first()
    if not lst:
        raise HTTPException(status_code=404, detail="List not found or not public")

    items = db.query(RankedListItem).filter(RankedListItem.list_id == list_id).order_by(asc(RankedListItem.rank)).all()
    owner = db.query(User).filter(User.id == friend_user_id).first()
    result = _list_dict(lst, db)
    result["items"] = [_item_dict(i) for i in items]
    result["owner"] = {"id": owner.id, "username": owner.username, "avatar_url": owner.avatar_url} if owner else None
    return result


# ── TMDB search proxy ─────────────────────────────────────────────────────────
@router.get("/search")
async def search_media(
    q: str = Query(..., min_length=1),
    media_type: str = Query("multi", regex="^(multi|movie|tv)$"),
):
    url = f"{TMDB_BASE}/search/{'multi' if media_type == 'multi' else media_type}"
    async with httpx.AsyncClient(timeout=10) as client:
        resp = await client.get(url, headers=TMDB_HEADERS, params={"query": q, "include_adult": False})
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail="TMDB search failed")
    data = resp.json()
    results = []
    for item in data.get("results", []):
        mt = item.get("media_type", media_type)
        if mt not in ("movie", "tv"):
            continue
        is_movie = mt == "movie"
        title = item.get("title") if is_movie else item.get("name")
        date_str = item.get("release_date") if is_movie else item.get("first_air_date")
        results.append({
            "id": str(item["id"]),
            "media_type": mt,
            "title": title,
            "poster_path": item.get("poster_path"),
            "release_year": date_str[:4] if date_str else None,
            "overview": item.get("overview"),
        })
    return {"results": results[:20]}


# ── Helpers ───────────────────────────────────────────────────────────────────
def _get_owned_list(list_id: int, user_id: int, db: Session) -> RankedList:
    lst = db.query(RankedList).filter(RankedList.id == list_id).first()
    if not lst:
        raise HTTPException(status_code=404, detail="List not found")
    if lst.user_id != user_id:
        raise HTTPException(status_code=403, detail="Not your list")
    return lst


def _list_dict(lst: RankedList, db: Session) -> dict:
    count = db.query(RankedListItem).filter(RankedListItem.list_id == lst.id).count()
    top_items = db.query(RankedListItem).filter(
        RankedListItem.list_id == lst.id,
        RankedListItem.poster_path.isnot(None),
    ).order_by(asc(RankedListItem.rank)).limit(4).all()
    return {
        "id": lst.id,
        "title": lst.title,
        "description": lst.description,
        "is_public": lst.is_public,
        "item_count": count,
        "cover_posters": [i.poster_path for i in top_items],
        "created_at": lst.created_at.isoformat() if lst.created_at else None,
        "updated_at": lst.updated_at.isoformat() if lst.updated_at else None,
    }


def _item_dict(i: RankedListItem) -> dict:
    return {
        "id": i.id, "list_id": i.list_id, "rank": i.rank,
        "media_id": i.media_id, "media_type": i.media_type, "title": i.title,
        "poster_path": i.poster_path, "release_year": i.release_year,
        "overview": i.overview, "note": i.note,
        "added_at": i.added_at.isoformat() if i.added_at else None,
    }


def _collection_dict(c: Collection, db: Session) -> dict:
    count = db.query(CollectionItem).filter(CollectionItem.collection_id == c.id).count()
    top_items = db.query(CollectionItem).filter(
        CollectionItem.collection_id == c.id,
        CollectionItem.poster_path.isnot(None),
    ).order_by(asc(CollectionItem.sort_order)).limit(4).all()
    return {
        "id": c.id, "title": c.title, "description": c.description,
        "cover_poster": c.cover_poster, "is_active": c.is_active,
        "item_count": count,
        "cover_posters": [i.poster_path for i in top_items],
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


def _col_item_dict(i: CollectionItem) -> dict:
    return {
        "id": i.id, "collection_id": i.collection_id, "sort_order": i.sort_order,
        "media_id": i.media_id, "media_type": i.media_type, "title": i.title,
        "poster_path": i.poster_path, "release_year": i.release_year, "overview": i.overview,
    }
